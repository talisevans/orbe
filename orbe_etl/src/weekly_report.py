"""
Weekly Performance Report Generator

Creates an Excel workbook with weekly performance summaries for each week of the year.
Each week runs Monday to Sunday.
"""

import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import os


class WeeklyReportGenerator:
    """
    Generates weekly performance reports in Excel format.
    """

    def __init__(self, detail_df, rebooking_df, productive_hours_df, clients_df, staff_df, logo_path=None):
        """
        Initialize the report generator with data.

        Args:
            detail_df: DataFrame with detailed sales/service transactions
            rebooking_df: DataFrame with rebooking information
            productive_hours_df: DataFrame with productive hours data
            logo_path: Optional path to the logo image file
        """
        self.detail_df = detail_df
        self.rebooking_df = rebooking_df
        self.productive_hours_df = productive_hours_df
        self.clients_df = clients_df
        self.staff_df = staff_df
        self.logo_path = logo_path
        self.workbook = Workbook()
        # Remove default sheet
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])

    def get_week_ranges(self, end_date=None):
        """
        Generate all week ranges (Monday to Sunday) from Jan 1 to the end of last complete week.

        Args:
            end_date: Optional end date (defaults to today)

        Returns:
            List of tuples: [(week_start, week_end, week_number), ...]
        """
        if end_date is None:
            end_date = datetime.now().date()

        # Find the last Sunday before or on end_date
        days_since_sunday = (end_date.weekday() + 1) % 7
        last_sunday = end_date - timedelta(days=days_since_sunday)

        # Start from Jan 1 of the current year
        year_start = datetime(end_date.year, 1, 1).date()

        # Find the first Monday of the year (or Jan 1 if it's a Monday)
        days_since_monday = year_start.weekday()
        if days_since_monday == 0:
            first_monday = year_start
        else:
            first_monday = year_start + timedelta(days=(7 - days_since_monday))

        weeks = []
        current_monday = first_monday
        week_number = 1

        while current_monday <= last_sunday:
            week_end = current_monday + timedelta(days=6)  # Sunday
            weeks.append((current_monday, week_end, week_number))
            current_monday += timedelta(days=7)
            week_number += 1

        return weeks

    def generate_all_weeks(self, end_date=None):
        """
        Generate a worksheet for each week in the year up to the last complete week.
        Weeks are generated in descending order (most recent first).

        Args:
            end_date: Optional end date (defaults to today)
        """
        weeks = self.get_week_ranges(end_date)

        # Reverse the weeks so most recent week is first
        for week_start, week_end, week_number in reversed(weeks):
            sheet_name = f"WE {week_end.strftime('%-d-%-m-%Y')}"
            self.create_week_worksheet(week_start, week_end, week_number, sheet_name)

    def create_week_worksheet(self, week_start, week_end, week_number, sheet_name):
        """
        Create a worksheet for a specific week with all formatting and data.

        Args:
            week_start: Start date of the week (Monday)
            week_end: End date of the week (Sunday)
            week_number: The week number in the year
            sheet_name: Name for the worksheet tab
        """
        ws = self.workbook.create_sheet(title=sheet_name)

        # Filter data for this week
        week_detail_df = self._filter_data_by_week(self.detail_df, week_start, week_end, 'Date')
        week_rebooking_df = self._filter_data_by_week(self.rebooking_df, week_start, week_end, 'AppointmentDate')
        week_productive_hours_df = self._filter_data_by_week(self.productive_hours_df, week_start, week_end, 'AppointmentDate')

        # Calculate report categories to determine column count
        report_categories = []
        if not week_detail_df.empty and 'ItemTypeStringCode' in week_detail_df.columns and 'ReportCategory' in week_detail_df.columns:
            report_categories = week_detail_df[week_detail_df['ItemTypeStringCode'] == 'ItemType.Service']['ReportCategory'].unique().tolist()
            report_categories = [item for item in report_categories if pd.notna(item)]

        # Calculate end column for background (report category columns + Total + spacer + productive hours + 1 extra)
        # Report category columns start at N (column 14, 1-indexed)
        service_start_col = 13  # Column N is index 13 (0-indexed)
        total_col_index = service_start_col + len(report_categories) + 1  # +1 for Total column
        # After service Total: spacer (+1), productive hours (+2), extra column (+3)
        end_col_for_bg = total_col_index + 3  # +3 for spacer, productive hours, and 1 extra

        # build the staff dataset for this week
        staff_dataset = self.buildStaffDataset(
            detail_df=week_detail_df,
            rebooking_df=week_rebooking_df,
            productive_hours_df=week_productive_hours_df,
            clients_df=self.clients_df,
            staff_df=self.staff_df,
            report_categories=report_categories
        )

        # Calculate unique customer counts for the week (to avoid double-counting clients who see multiple staff)
        unique_returning_clients = 0
        unique_new_clients = 0
        if not week_rebooking_df.empty and 'CustomerId' in week_rebooking_df.columns and 'is_new_client' in week_rebooking_df.columns:
            # Group by CustomerId to get unique customers
            customer_status = week_rebooking_df.groupby('CustomerId').agg({
                'is_new_client': 'max'  # If any appointment is new (True), customer is new
            }).reset_index()

            # Count unique returning and new clients
            unique_returning_clients = (customer_status['is_new_client'] == False).sum()
            unique_new_clients = (customer_status['is_new_client'] == True).sum()

        # Apply styling and structure
        self._setup_header_section(ws, week_number, week_start, week_end, end_col_for_bg)
        productive_hours_col_letter = self._setup_main_table_structure(ws, week_detail_df)
        total_row = self._populate_data(ws, staff_dataset, report_categories, productive_hours_col_letter, unique_returning_clients, unique_new_clients)
        self._add_summary_metrics(ws, total_row, staff_dataset, unique_returning_clients, unique_new_clients)
        self._apply_styling(ws)

        # Freeze panes: columns A-B and rows 1-8 remain visible when scrolling
        ws.freeze_panes = 'C9'

        return ws

    def _filter_data_by_week(self, df, week_start, week_end, date_column):
        """
        Filter dataframe to only include rows within the week range.

        Args:
            df: DataFrame to filter
            week_start: Start date of the week
            week_end: End date of the week
            date_column: Name of the date column to filter on

        Returns:
            Filtered DataFrame
        """
        if df is None or df.empty:
            return pd.DataFrame()

        df_copy = df.copy()

        # Ensure date column is in date format
        if date_column in df_copy.columns:
            df_copy[date_column] = pd.to_datetime(df_copy[date_column]).dt.date
            return df_copy[(df_copy[date_column] >= week_start) & (df_copy[date_column] <= week_end)]

        return pd.DataFrame()

    def _setup_header_section(self, ws, week_number, week_start, week_end, end_col_for_bg=30):
        """
        Setup the header section with title, week, month, and date range.

        Args:
            ws: Worksheet object
            week_number: The week number
            week_start: Start date of the week
            week_end: End date of the week
            end_col_for_bg: Number of columns to apply background to (default 30)
        """
        # Apply medium gray background to entire sheet first
        for row in range(1, 100):  # Apply to first 100 rows
            for col in range(1, end_col_for_bg + 1):  # Apply to dynamic column count
                cell = ws.cell(row=row, column=col)
                cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

        # Row 1: Gray background (thin spacer bar)
        ws.row_dimensions[1].height = 15

        # Row 2: Title row "Orbe North Adelaide"
        ws.row_dimensions[2].height = 35
        ws['B2'] = 'Orbe North Adelaide'
        ws['B2'].font = Font(color="FFFFFF", bold=True, size=14)
        ws['B2'].fill = PatternFill(start_color="3A3838", end_color="3A3838", fill_type="solid")
        ws['B2'].alignment = Alignment(horizontal='left', vertical='center')

        # Row 3: Gray background (spacer bar)
        ws.row_dimensions[3].height = 15

        # Row 4: Header row with ORBE logo, week, month, and dates
        ws.row_dimensions[4].height = 35

        # White border for creating boxes (thinner borders)
        white_border = Border(
            left=Side(style='thin', color='FFFFFF'),
            right=Side(style='thin', color='FFFFFF'),
            top=Side(style='thin', color='FFFFFF'),
            bottom=Side(style='thin', color='FFFFFF')
        )

        # Insert logo image
        logo_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'img', 'orbe-logo.png')
        if os.path.exists(logo_path):
            img = Image(logo_path)
            # Scale the image to fit nicely in the merged cell (adjust size as needed)
            img.width = 130  # Adjust width in pixels
            img.height = 40   # Adjust height in pixels
            ws.add_image(img, 'B4')

        # set the background colour of B4 to white
        ws['B4'].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        ws['B4'].border = white_border

        # Week label (C4, with dark background #3A3838)
        ws['C4'] = 'Week'
        ws['C4'].font = Font(color="FFFFFF", bold=True, size=11)
        ws['C4'].fill = PatternFill(start_color="3A3838", end_color="3A3838", fill_type="solid")
        ws['C4'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C4'].border = white_border

        # Merge D & E for week number
        ws.merge_cells('D4:E4')

        # Week number value (D4, with light background #D9D9D9)
        ws['D4'] = f'Week {week_number}'
        ws['D4'].font = Font(size=11)
        ws['D4'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        ws['D4'].alignment = Alignment(horizontal='center', vertical='center')
        ws['D4'].border = white_border
        ws['E4'].border = white_border  # Apply border to merged cell

        # Month label (G4, with dark background #3A3838)
        ws['G4'] = 'Month'
        ws['G4'].font = Font(color="FFFFFF", bold=True, size=11)
        ws['G4'].fill = PatternFill(start_color="3A3838", end_color="3A3838", fill_type="solid")
        ws['G4'].alignment = Alignment(horizontal='center', vertical='center')
        ws['G4'].border = white_border

        # Month name (H4, with light background #D9D9D9)
        ws['H4'] = week_start.strftime('%b')
        ws['H4'].font = Font(size=11)
        ws['H4'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        ws['H4'].alignment = Alignment(horizontal='center', vertical='center')
        ws['H4'].border = white_border

        # Start Date label (K4, with dark background #3A3838)
        ws['K4'] = 'Start Date'
        ws['K4'].font = Font(color="FFFFFF", bold=True, size=10)
        ws['K4'].fill = PatternFill(start_color="3A3838", end_color="3A3838", fill_type="solid")
        ws['K4'].alignment = Alignment(horizontal='center', vertical='center')
        ws['K4'].border = white_border

        # Start date value (L4, with light background #D9D9D9)
        ws['L4'] = week_start
        ws['L4'].font = Font(size=11)
        ws['L4'].number_format = 'D-MMM-YY'
        ws['L4'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        ws['L4'].alignment = Alignment(horizontal='center', vertical='center')
        ws['L4'].border = white_border

        # End Date label (N4, with dark background #3A3838)
        ws['N4'] = 'End Date'
        ws['N4'].font = Font(color="FFFFFF", bold=True, size=10)
        ws['N4'].fill = PatternFill(start_color="3A3838", end_color="3A3838", fill_type="solid")
        ws['N4'].alignment = Alignment(horizontal='center', vertical='center')
        ws['N4'].border = white_border

        # End date value (O4, with light background #D9D9D9)
        ws['O4'] = week_end
        ws['O4'].font = Font(size=11)
        ws['O4'].number_format = 'D-MMM-YY'
        ws['O4'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        ws['O4'].alignment = Alignment(horizontal='center', vertical='center')
        ws['O4'].border = white_border

    def _setup_main_table_structure(self, ws, week_detail_df):
        """
        Setup the main table structure with headers and merged cells.

        Args:
            ws: Worksheet object
        """
        # White border for creating boxes
        white_border = Border(
            left=Side(style='thin', color='FFFFFF'),
            right=Side(style='thin', color='FFFFFF'),
            top=Side(style='thin', color='FFFFFF'),
            bottom=Side(style='thin', color='FFFFFF')
        )

        # Row 5: Spacer row
        ws.row_dimensions[5].height = 15

        # Row 6: Spacer row
        ws.row_dimensions[6].height = 15

        # Row 7: Section headers with SPACER columns A, F, I, M
        ws.row_dimensions[7].height = 30

        # Spacer columns in row 7
        for col in ['A', 'F', 'I', 'M']:
            ws[f'{col}7'].fill = PatternFill(start_color="757171", end_color="757171", fill_type="solid")

        # Staff + No. of Clients (C7:E7 merged)
        ws.merge_cells('C7:E7')
        ws['C7'] = 'No. of Clients'
        ws['C7'].font = Font(color="FFFFFF", bold=True, size=11)
        ws['C7'].fill = PatternFill(start_color="3A3838", end_color="3A3838", fill_type="solid")
        ws['C7'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C7'].border = white_border
        ws['D7'].border = white_border
        ws['E7'].border = white_border

        # ($) Sales (G7:H7 merged)
        ws.merge_cells('G7:H7')
        ws['G7'] = '($) Sales'
        ws['G7'].font = Font(color="FFFFFF", bold=True, size=11)
        ws['G7'].fill = PatternFill(start_color="3A3838", end_color="3A3838", fill_type="solid")
        ws['G7'].alignment = Alignment(horizontal='center', vertical='center')
        ws['G7'].border = white_border
        ws['H7'].border = white_border

        # (%) Rebooked (J7:L7 merged)
        ws.merge_cells('J7:L7')
        ws['J7'] = '(%) Rebooked'
        ws['J7'].font = Font(color="FFFFFF", bold=True, size=11)
        ws['J7'].fill = PatternFill(start_color="3A3838", end_color="3A3838", fill_type="solid")
        ws['J7'].alignment = Alignment(horizontal='center', vertical='center')
        ws['J7'].border = white_border
        ws['K7'].border = white_border
        ws['L7'].border = white_border

        # get a unique list of ReportCategory from week_detail_df where `ItemTypeStringCode` is 'ItemType.Service'
        report_categories = []
        if not week_detail_df.empty and 'ItemTypeStringCode' in week_detail_df.columns and 'ReportCategory' in week_detail_df.columns:
            report_categories = week_detail_df[week_detail_df['ItemTypeStringCode'] == 'ItemType.Service']['ReportCategory'].unique().tolist()
            report_categories = [item for item in report_categories if pd.notna(item)]

        # (#) Service header - dynamically merge based on number of report categories
        # Service columns start at N (column 14) and include all report categories plus Total column
        service_start_col = 13  # Column N is index 13 (0-indexed)
        total_col_index = service_start_col + len(report_categories) + 1
        total_col_letter = get_column_letter(total_col_index)
        ws.merge_cells(f'N7:{total_col_letter}7')
        ws['N7'] = '(#) Service'
        ws['N7'].font = Font(color="FFFFFF", bold=True, size=11)
        ws['N7'].fill = PatternFill(start_color="3A3838", end_color="3A3838", fill_type="solid")
        ws['N7'].alignment = Alignment(horizontal='center', vertical='center')
        ws['N7'].border = white_border

        # Row 8: Column headers with SPACER columns A, F, I, M
        ws.row_dimensions[8].height = 35

        # Spacer columns in row 8
        for col in ['A', 'F', 'I', 'M']:
            ws[f'{col}8'].fill = PatternFill(start_color="757171", end_color="757171", fill_type="solid")

        # Fixed headers (non-service columns)
        headers = {
            'B8': 'Staff',
            'C8': 'Reg',
            'D8': 'New',
            'E8': 'Total',
            'G8': '($) Total',
            'H8': '($) Avg',
            'J8': 'Reg.',
            'K8': 'New',
            'L8': 'Total',
        }

        for cell, value in headers.items():
            ws[cell] = value
            ws[cell].font = Font(color="FFFFFF", bold=True, size=9)
            ws[cell].fill = PatternFill(start_color="3A3838", end_color="3A3838", fill_type="solid")
            ws[cell].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws[cell].border = white_border

        # Dynamic report category headers starting from column N
        service_start_col = 13  # Column N is index 13 (0-indexed)
        for i, category_name in enumerate(report_categories):
            col_letter = get_column_letter(service_start_col + i + 1)  # +1 for 1-indexed
            cell = f'{col_letter}8'
            ws[cell] = category_name
            ws[cell].font = Font(color="FFFFFF", bold=True, size=9)
            ws[cell].fill = PatternFill(start_color="3A3838", end_color="3A3838", fill_type="solid")
            ws[cell].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws[cell].border = white_border
            ws[f'{col_letter}7'].border = white_border  # Also apply border to merged header cell

        # Total column comes after all report categories
        total_col_index = service_start_col + len(report_categories) + 1
        total_col_letter = get_column_letter(total_col_index)
        ws[f'{total_col_letter}8'] = 'Total'
        ws[f'{total_col_letter}8'].font = Font(color="FFFFFF", bold=True, size=9)
        ws[f'{total_col_letter}8'].fill = PatternFill(start_color="3A3838", end_color="3A3838", fill_type="solid")
        ws[f'{total_col_letter}8'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws[f'{total_col_letter}8'].border = white_border
        ws[f'{total_col_letter}7'].border = white_border  # Also apply border to merged header cell

        # (#) Productive Hours section comes after service Total
        # Add spacer column first
        spacer_col_index = total_col_index + 1
        spacer_col_letter = get_column_letter(spacer_col_index)
        ws[f'{spacer_col_letter}7'].fill = PatternFill(start_color="757171", end_color="757171", fill_type="solid")
        ws[f'{spacer_col_letter}8'].fill = PatternFill(start_color="757171", end_color="757171", fill_type="solid")

        # Productive Hours column
        productive_hours_col_index = spacer_col_index + 1
        productive_hours_col_letter = get_column_letter(productive_hours_col_index)

        # Row 7: Section header
        ws[f'{productive_hours_col_letter}7'] = 'Prd. Hrs'
        ws[f'{productive_hours_col_letter}7'].font = Font(color="FFFFFF", bold=True, size=11)
        ws[f'{productive_hours_col_letter}7'].fill = PatternFill(start_color="3A3838", end_color="3A3838", fill_type="solid")
        ws[f'{productive_hours_col_letter}7'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws[f'{productive_hours_col_letter}7'].border = white_border

        # Row 8: Column header
        ws[f'{productive_hours_col_letter}8'] = 'Hours'
        ws[f'{productive_hours_col_letter}8'].font = Font(color="FFFFFF", bold=True, size=9)
        ws[f'{productive_hours_col_letter}8'].fill = PatternFill(start_color="3A3838", end_color="3A3838", fill_type="solid")
        ws[f'{productive_hours_col_letter}8'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws[f'{productive_hours_col_letter}8'].border = white_border

        # Set column widths with spacers
        column_widths = {
            'A': 2, 'B': 21, 'C': 10, 'D': 10, 'E': 10,
            'F': 2, 'G': 10, 'H': 10,
            'I': 2, 'J': 10, 'K': 10, 'L': 10,
            'M': 2
        }

        # Set width for productive hours spacer and column
        column_widths[spacer_col_letter] = 2
        column_widths[productive_hours_col_letter] = 10

        # Add dynamic report category column widths
        for i in range(len(report_categories) + 1):  # +1 for Total column
            col_letter = get_column_letter(service_start_col + i + 1)
            column_widths[col_letter] = 10

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        return productive_hours_col_letter

    def _populate_data(self, ws, staff_dataset, report_categories, productive_hours_col_letter, unique_returning_clients, unique_new_clients):
        """
        Populate the worksheet with calculated data from the staff dataset.

        Args:
            ws: Worksheet object
            staff_dataset: DataFrame with staff data for the week
            report_categories: List of report categories for this week
            productive_hours_col_letter: Column letter for productive hours
            unique_returning_clients: Unique count of returning clients for the week (to avoid double-counting)
            unique_new_clients: Unique count of new clients for the week (to avoid double-counting)
        """
        start_row = 9  # Data starts at row 9
        current_row = start_row

        # White border for data cells
        white_border = Border(
            left=Side(style='thin', color='FFFFFF'),
            right=Side(style='thin', color='FFFFFF'),
            top=Side(style='thin', color='FFFFFF'),
            bottom=Side(style='thin', color='FFFFFF')
        )

        # sort staff_dataset by StaffName if column exists
        if 'StaffName' in staff_dataset.columns:
            staff_dataset = staff_dataset.sort_values(by='StaffName')

        # Iterate through staff_dataset
        row_index = 0  # Track row for alternating colors
        for index, staff_row in staff_dataset.iterrows():

            # Determine if this is an even row (for alternating colors)
            is_even_row = (row_index % 2 == 1)
            alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") if is_even_row else None

            # Spacer columns (A, F, I, M) always have medium gray
            for col in ['A', 'F', 'I', 'M']:
                ws[f'{col}{current_row}'].fill = PatternFill(start_color="757171", end_color="757171", fill_type="solid")

            # B: Staff name (first name only, uppercase)
            staff_name = staff_row['StaffName'] if 'StaffName' in staff_row else ''
            # Handle NaN or float values
            if pd.isna(staff_name):
                staff_name = ''
            ws[f'B{current_row}'] = str(staff_name).upper()
            ws[f'B{current_row}'].border = white_border
            ws[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
            if alt_fill:
                ws[f'B{current_row}'].fill = alt_fill
            else:
                ws[f'B{current_row}'].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

            # No. of Clients (C, D, E)
            returning_clients_val = staff_row.get('returningClients', 0)
            ws[f'C{current_row}'] = '-' if returning_clients_val == 0 else returning_clients_val  # Reg clients
            ws[f'C{current_row}'].border = white_border
            ws[f'C{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            if alt_fill:
                ws[f'C{current_row}'].fill = alt_fill
            else:
                ws[f'C{current_row}'].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

            new_clients_val = staff_row.get('newClients', 0)
            ws[f'D{current_row}'] = '-' if new_clients_val == 0 else new_clients_val  # New clients
            ws[f'D{current_row}'].border = white_border
            ws[f'D{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            if alt_fill:
                ws[f'D{current_row}'].fill = alt_fill
            else:
                ws[f'D{current_row}'].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

            total_clients_val = staff_row.get('totalClients', 0)
            ws[f'E{current_row}'] = '-' if total_clients_val == 0 else total_clients_val  # Total clients
            ws[f'E{current_row}'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            ws[f'E{current_row}'].border = white_border
            ws[f'E{current_row}'].alignment = Alignment(horizontal='center', vertical='center')

            # Sales (G, H)
            total_sales_val = staff_row.get('totalSales', 0.0)
            ws[f'G{current_row}'] = '-' if total_sales_val == 0 else total_sales_val  # ($) Total
            if total_sales_val != 0:
                ws[f'G{current_row}'].number_format = '$#,##0'
            ws[f'G{current_row}'].border = white_border
            ws[f'G{current_row}'].alignment = Alignment(horizontal='right', vertical='center')
            if alt_fill:
                ws[f'G{current_row}'].fill = alt_fill
            else:
                ws[f'G{current_row}'].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

            avg_sales_val = staff_row.get('avgSales', 0.0)
            ws[f'H{current_row}'] = '-' if avg_sales_val == 0 else avg_sales_val  # ($) Avg
            if avg_sales_val != 0:
                ws[f'H{current_row}'].number_format = '$#,##0'
            ws[f'H{current_row}'].border = white_border
            ws[f'H{current_row}'].alignment = Alignment(horizontal='right', vertical='center')
            if alt_fill:
                ws[f'H{current_row}'].fill = alt_fill
            else:
                ws[f'H{current_row}'].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

            # (%) Rebooked (J, K)
            # Calculate rebooking percentage for returning clients
            returning_clients = staff_row.get('returningClients', 0)
            rebooked_returning = staff_row.get('rebookedReturning', 0)
            rebooking_pct_returning = rebooked_returning / returning_clients if returning_clients > 0 else 0
            ws[f'J{current_row}'] = '-' if rebooking_pct_returning == 0 else rebooking_pct_returning  # Reg.
            if rebooking_pct_returning != 0:
                ws[f'J{current_row}'].number_format = '0%'
            ws[f'J{current_row}'].border = white_border
            ws[f'J{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            if alt_fill:
                ws[f'J{current_row}'].fill = alt_fill
            else:
                ws[f'J{current_row}'].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

            # Calculate rebooking percentage for new clients
            new_clients = staff_row.get('newClients', 0)
            rebooked_new = staff_row.get('rebookedNew', 0)
            rebooking_pct_new = rebooked_new / new_clients if new_clients > 0 else 0
            ws[f'K{current_row}'] = '-' if rebooking_pct_new == 0 else rebooking_pct_new  # New
            if rebooking_pct_new != 0:
                ws[f'K{current_row}'].number_format = '0%'
            ws[f'K{current_row}'].border = white_border
            ws[f'K{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            if alt_fill:
                ws[f'K{current_row}'].fill = alt_fill
            else:
                ws[f'K{current_row}'].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

            # Calculate total rebooking percentage (all clients)
            total_clients = staff_row.get('totalClients', 0)
            total_rebooked = staff_row.get('rebookedReturning', 0) + staff_row.get('rebookedNew', 0)
            rebooking_pct_total = total_rebooked / total_clients if total_clients > 0 else 0
            ws[f'L{current_row}'] = '-' if rebooking_pct_total == 0 else rebooking_pct_total  # Total
            if rebooking_pct_total != 0:
                ws[f'L{current_row}'].number_format = '0%'
            ws[f'L{current_row}'].border = white_border
            ws[f'L{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'L{current_row}'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

            # (#) Service counts - dynamically loop through report categories
            service_start_col = 13  # Column N is index 13 (0-indexed)
            category_counts = staff_row.get('categoryCounts', {})
            total_services = 0

            for i, category in enumerate(report_categories):
                col_letter = get_column_letter(service_start_col + i + 1)
                category_count = category_counts.get(category, 0)
                ws[f'{col_letter}{current_row}'] = '-' if category_count == 0 else category_count
                ws[f'{col_letter}{current_row}'].border = white_border
                ws[f'{col_letter}{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
                if alt_fill:
                    ws[f'{col_letter}{current_row}'].fill = alt_fill
                else:
                    ws[f'{col_letter}{current_row}'].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                total_services += category_count

            # Total services column (comes after all report categories)
            total_col_index = service_start_col + len(report_categories) + 1
            total_col_letter = get_column_letter(total_col_index)
            ws[f'{total_col_letter}{current_row}'] = '-' if total_services == 0 else total_services  # Total services
            ws[f'{total_col_letter}{current_row}'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            ws[f'{total_col_letter}{current_row}'].border = white_border
            ws[f'{total_col_letter}{current_row}'].alignment = Alignment(horizontal='center', vertical='center')

            # Spacer column before productive hours
            spacer_col_index = total_col_index + 1
            spacer_col_letter = get_column_letter(spacer_col_index)
            ws[f'{spacer_col_letter}{current_row}'].fill = PatternFill(start_color="757171", end_color="757171", fill_type="solid")

            # (#) Productive Hours
            productive_hours_val = staff_row.get('productiveHours', 0.0)
            ws[f'{productive_hours_col_letter}{current_row}'] = '-' if productive_hours_val == 0 else productive_hours_val
            if productive_hours_val != 0:
                ws[f'{productive_hours_col_letter}{current_row}'].number_format = '0.0'
            ws[f'{productive_hours_col_letter}{current_row}'].border = white_border
            ws[f'{productive_hours_col_letter}{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            if alt_fill:
                ws[f'{productive_hours_col_letter}{current_row}'].fill = alt_fill
            else:
                ws[f'{productive_hours_col_letter}{current_row}'].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

            # set the row height to 26
            ws.row_dimensions[current_row].height = 26

            current_row += 1
            row_index += 1

        # Add TOTAL row with formulas
        total_row = current_row

        # Spacer columns (A, F, I, M) in TOTAL row
        for col in ['A', 'F', 'I', 'M']:
            ws[f'{col}{total_row}'].fill = PatternFill(start_color="757171", end_color="757171", fill_type="solid")

        ws[f'B{total_row}'] = 'TOTAL'
        ws[f'B{total_row}'].font = Font(bold=True)
        ws[f'B{total_row}'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        ws[f'B{total_row}'].border = white_border
        ws[f'B{total_row}'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[total_row].height = 26

        # Add TOTAL row values (using unique client counts to avoid double-counting)
        if current_row > start_row:  # Only add formulas if there's data
            # Calculate unique total clients
            unique_total_clients = unique_returning_clients + unique_new_clients

            # Use unique counts for client columns (not SUM formulas to avoid double-counting)
            # For sales, use SUM formula
            formulas = {
                'C': unique_returning_clients,  # Unique reg clients
                'D': unique_new_clients,  # Unique new clients
                'E': unique_total_clients,  # Unique total clients
                'G': f'=SUM(G{start_row}:G{current_row - 1})',  # ($) Total
                'H': f'=G{total_row}/E{total_row}',  # ($) Avg - total sales / total clients
            }

            for col, formula in formulas.items():
                ws[f'{col}{total_row}'] = formula
                ws[f'{col}{total_row}'].font = Font(bold=True)
                ws[f'{col}{total_row}'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                ws[f'{col}{total_row}'].border = white_border
                # Right align sales columns (G, H), center align others
                if col in ['G', 'H']:
                    ws[f'{col}{total_row}'].alignment = Alignment(horizontal='right', vertical='center')
                    ws[f'{col}{total_row}'].number_format = '$#,##0'
                else:
                    ws[f'{col}{total_row}'].alignment = Alignment(horizontal='center', vertical='center')

            # Calculate rebooking percentages from absolute values in staff_dataset
            # Sum the underlying counts across all staff
            total_returning_clients = staff_dataset['returningClients'].sum()
            total_rebooked_returning = staff_dataset['rebookedReturning'].sum()
            total_new_clients = staff_dataset['newClients'].sum()
            total_rebooked_new = staff_dataset['rebookedNew'].sum()

            # Calculate percentages
            rebooking_pct_returning_total = total_rebooked_returning / total_returning_clients if total_returning_clients > 0 else 0
            rebooking_pct_new_total = total_rebooked_new / total_new_clients if total_new_clients > 0 else 0

            # Populate rebooking percentages (J, K)
            ws[f'J{total_row}'] = rebooking_pct_returning_total
            ws[f'J{total_row}'].font = Font(bold=True)
            ws[f'J{total_row}'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            ws[f'J{total_row}'].border = white_border
            ws[f'J{total_row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'J{total_row}'].number_format = '0%'

            ws[f'K{total_row}'] = rebooking_pct_new_total
            ws[f'K{total_row}'].font = Font(bold=True)
            ws[f'K{total_row}'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            ws[f'K{total_row}'].border = white_border
            ws[f'K{total_row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'K{total_row}'].number_format = '0%'

            # Calculate total rebooking percentage (all clients combined)
            total_all_clients = total_returning_clients + total_new_clients
            total_all_rebooked = total_rebooked_returning + total_rebooked_new
            rebooking_pct_all_total = total_all_rebooked / total_all_clients if total_all_clients > 0 else 0

            ws[f'L{total_row}'] = rebooking_pct_all_total
            ws[f'L{total_row}'].font = Font(bold=True)
            ws[f'L{total_row}'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            ws[f'L{total_row}'].border = white_border
            ws[f'L{total_row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'L{total_row}'].number_format = '0%'

            # Add formulas for dynamic report category columns
            service_start_col = 13  # Column N is index 13 (0-indexed)
            for i in range(len(report_categories)):
                col_letter = get_column_letter(service_start_col + i + 1)
                ws[f'{col_letter}{total_row}'] = f'=SUM({col_letter}{start_row}:{col_letter}{current_row - 1})'
                ws[f'{col_letter}{total_row}'].font = Font(bold=True)
                ws[f'{col_letter}{total_row}'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                ws[f'{col_letter}{total_row}'].border = white_border
                ws[f'{col_letter}{total_row}'].alignment = Alignment(horizontal='center', vertical='center')

            # Total services column formula
            total_col_index = service_start_col + len(report_categories) + 1
            total_col_letter = get_column_letter(total_col_index)
            ws[f'{total_col_letter}{total_row}'] = f'=SUM({total_col_letter}{start_row}:{total_col_letter}{current_row - 1})'
            ws[f'{total_col_letter}{total_row}'].font = Font(bold=True)
            ws[f'{total_col_letter}{total_row}'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            ws[f'{total_col_letter}{total_row}'].border = white_border
            ws[f'{total_col_letter}{total_row}'].alignment = Alignment(horizontal='center', vertical='center')

            # Spacer column before productive hours in TOTAL row
            spacer_col_index = total_col_index + 1
            spacer_col_letter = get_column_letter(spacer_col_index)
            ws[f'{spacer_col_letter}{total_row}'].fill = PatternFill(start_color="757171", end_color="757171", fill_type="solid")

            # Productive Hours total
            ws[f'{productive_hours_col_letter}{total_row}'] = f'=SUM({productive_hours_col_letter}{start_row}:{productive_hours_col_letter}{current_row - 1})'
            ws[f'{productive_hours_col_letter}{total_row}'].font = Font(bold=True)
            ws[f'{productive_hours_col_letter}{total_row}'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            ws[f'{productive_hours_col_letter}{total_row}'].border = white_border
            ws[f'{productive_hours_col_letter}{total_row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'{productive_hours_col_letter}{total_row}'].number_format = '0.0'

        return total_row  # Return the total row for use in summary section

    def buildStaffDataset(self, detail_df, rebooking_df, productive_hours_df, clients_df, staff_df, report_categories):
        """
        Build the staff dataset for the week.

        Args:
            detail_df: Filtered detail dataframe for the week
            rebooking_df: Filtered rebooking dataframe for the week
            productive_hours_df: Filtered productive hours dataframe for the week
            clients_df: Clients dataframe
            staff_df: Staff dataframe
            report_categories: List of report categories for this week

        Returns:
            DataFrame with staff dataset for the week
        """

        # declare the dataset to return
        dataset = []

        # get a unique list of `employeeId` from detail_df
        staff_ids = []
        if not detail_df.empty and 'EmployeeId' in detail_df.columns:
            staff_ids = detail_df['EmployeeId'].dropna().unique().tolist()
        
        # get a unique list of EmployeeId from the rebooking_df and add to staff_ids if not already present
        if not rebooking_df.empty and 'EmployeeId' in rebooking_df.columns:
            rebooking_staff_ids = rebooking_df['EmployeeId'].dropna().unique().tolist()
            for staff_id in rebooking_staff_ids:
                if staff_id not in staff_ids:
                    staff_ids.append(staff_id)

    
        # Loop through each staff member to build their data
        for staff_id in staff_ids:

            # find this staffId in the staff_df to get their name
            staff_info = staff_df[staff_df['EmployeeId'] == staff_id] if not staff_df.empty else pd.DataFrame()

            # get the FirstName and Surname (if available)
            full_name = ''
            if not staff_info.empty:
                first_name = staff_info['FirstName'].iloc[0] if 'FirstName' in staff_info.columns else ''
                surname = staff_info['Surname'].iloc[0] if 'Surname' in staff_info.columns else ''

                # capitalize the first letter of first name and surname, the rest lowercase
                first_name = first_name.capitalize() if isinstance(first_name, str) else ''
                surname = surname.capitalize() if isinstance(surname, str) else ''
                full_name = f"{first_name} {surname}".strip()

            # filter data for this staff member
            staff_rebooking = rebooking_df[rebooking_df['EmployeeId'] == staff_id] if not rebooking_df.empty else pd.DataFrame()
            staff_detail_df = detail_df[detail_df['EmployeeId'] == staff_id] if not detail_df.empty else pd.DataFrame()

            # declare the dataset entry for this staff member
            staff_data = {
                'EmployeeId': staff_id,
                'StaffName': full_name,

                'returningClients': 0,  # Placeholder
                'rebookedReturning': 0.0,  # Placeholder

                'newClients': 0,        # Placeholder
                'rebookedNew': 0.0,  # Placeholder

                'totalClients': 0,      # Placeholder

                'totalSales': 0.0,      # Placeholder
                'avgSales': 0.0,        # Placeholder

                'productiveHours': 0.0,   # Placeholder

                'categoryCounts': {category: 0 for category in report_categories}  # Placeholder
            }

            # Determine client status for each unique customer
            if not staff_rebooking.empty and 'CustomerId' in staff_rebooking.columns and 'is_new_client' in staff_rebooking.columns:
                # Group by CustomerId and aggregate
                # A customer is "new" if ANY of their appointments in the week is marked as new
                # A customer is "rebooked" if ANY of their appointments show they rebooked
                customer_status = staff_rebooking.groupby('CustomerId').agg({
                    'is_new_client': 'max',  # If any appointment is new (True), customer is new
                    'is_rebooked': 'max'     # If any appointment shows rebooking, customer rebooked
                }).reset_index()

                # Count returning clients (not new)
                returning_clients = (customer_status['is_new_client'] == False).sum()
                staff_data['returningClients'] = returning_clients

                # Count returning clients who rebooked
                rebooked_returning = ((customer_status['is_new_client'] == False) &
                                     (customer_status['is_rebooked'] == True)).sum()
                staff_data['rebookedReturning'] = rebooked_returning

                # Count new clients
                new_clients = (customer_status['is_new_client'] == True).sum()
                staff_data['newClients'] = new_clients

                # Count new clients who rebooked
                rebooked_new = ((customer_status['is_new_client'] == True) &
                               (customer_status['is_rebooked'] == True)).sum()
                staff_data['rebookedNew'] = rebooked_new

            # sum the `LineExTaxAmount` from staff_detail_df for totalSales
            if not staff_detail_df.empty and 'LineExTaxAmount' in staff_detail_df.columns:
                total_sales = staff_detail_df['LineExTaxAmount'].sum()
                staff_data['totalSales'] = total_sales

            # calculate avgSales as totalSales divided by totalClients
            total_clients = staff_data['returningClients'] + staff_data['newClients']
            staff_data['totalClients'] = total_clients
            if total_clients > 0:
                staff_data['avgSales'] = staff_data['totalSales'] / total_clients

            # loop through the report_categories and count occurrences in staff_detail_df
            for category in report_categories:
                if not staff_detail_df.empty and 'ReportCategory' in staff_detail_df.columns:
                    category_count = staff_detail_df[staff_detail_df['ReportCategory'] == category].shape[0]
                    staff_data['categoryCounts'][category] = category_count

            # if we have this EmployeeId in productive_hours_df, sum their DurationMinutes
            if not productive_hours_df.empty and 'EmployeeId' in productive_hours_df.columns:

                staff_hours = productive_hours_df[productive_hours_df['EmployeeId'] == staff_id].copy()

                # set duration minutes to a number
                staff_hours['DurationMinutes'] = pd.to_numeric(staff_hours['DurationMinutes'], errors='coerce').fillna(0)
                total_minutes = staff_hours['DurationMinutes'].sum()
                staff_data['productiveHours'] = total_minutes / 60  # Convert to hours

            # append the staff_data to dataset
            dataset.append(staff_data)

        # convert dataset to DataFrame
        dataset_df = pd.DataFrame(dataset)

        return dataset_df

    def _add_summary_metrics(self, ws, total_row, staff_dataset, unique_returning_clients, unique_new_clients):
        """
        Add summary metrics section at the bottom (Total Clients, Rebooking %, Chemical %).

        Args:
            ws: Worksheet object
            total_row: The row number where the TOTAL row is located
            staff_dataset: DataFrame with staff data for calculating rebooking percentage
            unique_returning_clients: Unique count of returning clients for the week
            unique_new_clients: Unique count of new clients for the week
        """
        # Summary section starts 2 rows below the TOTAL row
        summary_start_row = total_row + 2

        # White border for summary boxes
        white_border = Border(
            left=Side(style='medium', color='FFFFFF'),
            right=Side(style='medium', color='FFFFFF'),
            top=Side(style='medium', color='FFFFFF'),
            bottom=Side(style='medium', color='FFFFFF')
        )

        # Row 1: # Total Clients
        ws[f'B{summary_start_row}'] = '# Total Clients'
        ws[f'B{summary_start_row}'].font = Font(color="FFFFFF", size=11)
        ws[f'B{summary_start_row}'].fill = PatternFill(start_color="3A3838", end_color="3A3838", fill_type="solid")
        ws[f'B{summary_start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'B{summary_start_row}'].border = white_border
        ws.row_dimensions[summary_start_row].height = 26

        # Reference the TOTAL row column E for total clients
        ws[f'C{summary_start_row}'] = f'=E{total_row}'
        ws[f'C{summary_start_row}'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        ws[f'C{summary_start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'C{summary_start_row}'].border = white_border
        ws[f'C{summary_start_row}'].font = Font(bold=True)

        # Row 2: (%) Rebooking
        ws[f'B{summary_start_row + 1}'] = '(%) Rebooking'
        ws[f'B{summary_start_row + 1}'].font = Font(color="FFFFFF", size=11)
        ws[f'B{summary_start_row + 1}'].fill = PatternFill(start_color="3A3838", end_color="3A3838", fill_type="solid")
        ws[f'B{summary_start_row + 1}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'B{summary_start_row + 1}'].border = white_border
        ws.row_dimensions[summary_start_row + 1].height = 26

        # Calculate total rebooking percentage from staff_dataset (same as TOTAL row)
        if not staff_dataset.empty and 'returningClients' in staff_dataset.columns:
            total_returning_clients = staff_dataset['returningClients'].sum()
            total_rebooked_returning = staff_dataset['rebookedReturning'].sum()
            total_new_clients = staff_dataset['newClients'].sum()
            total_rebooked_new = staff_dataset['rebookedNew'].sum()
            total_all_clients = total_returning_clients + total_new_clients
            total_all_rebooked = total_rebooked_returning + total_rebooked_new
            rebooking_pct = total_all_rebooked / total_all_clients if total_all_clients > 0 else 0
        else:
            rebooking_pct = 0

        # calculate percent chemical (using unique client count)
        unique_total_clients = unique_returning_clients + unique_new_clients
        percenet_chemical = self._calcPercentChemical(staff_dataset, unique_total_clients)

        ws[f'C{summary_start_row + 1}'] = rebooking_pct
        ws[f'C{summary_start_row + 1}'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        ws[f'C{summary_start_row + 1}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'C{summary_start_row + 1}'].number_format = '0%'  # Format as percentage
        ws[f'C{summary_start_row + 1}'].border = white_border
        ws[f'C{summary_start_row + 1}'].font = Font(bold=True)

        # Row 3: (%) Chemical
        ws[f'B{summary_start_row + 2}'] = '(%) Chemical'
        ws[f'B{summary_start_row + 2}'].font = Font(color="FFFFFF", size=11)
        ws[f'B{summary_start_row + 2}'].fill = PatternFill(start_color="3A3838", end_color="3A3838", fill_type="solid")
        ws[f'B{summary_start_row + 2}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'B{summary_start_row + 2}'].border = white_border
        ws.row_dimensions[summary_start_row + 2].height = 26

        # Chemical % remains blank
        ws[f'C{summary_start_row + 2}'] = percenet_chemical
        ws[f'C{summary_start_row + 2}'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        ws[f'C{summary_start_row + 2}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'C{summary_start_row + 2}'].number_format = '0%'  # Format as percentage
        ws[f'C{summary_start_row + 2}'].border = white_border
        ws[f'C{summary_start_row + 2}'].font = Font(bold=True)

    def _apply_styling(self, ws):
        """
        Apply final styling touches to the worksheet.

        Args:
            ws: Worksheet object
        """
        # No additional styling needed - all styling is applied directly to cells
        pass

    def _calcPercentChemical(self, staff_dataset, unique_total_clients):
        """
        Calculate the percentage of chemical services.

        Args:
            staff_dataset: DataFrame with staff data for calculating chemical percentage
            unique_total_clients: Unique count of total clients (to avoid double-counting)

        Returns:
            Percentage of chemical services as a float
        """

        # use the provided unique total clients count
        total_clients = unique_total_clients

        # declare total colours
        total_colours = 0

        # loop through the staff_dataset to sum colour services
        for index, staff_row in staff_dataset.iterrows():
            category_counts = staff_row.get('categoryCounts', {})
            colour_count = category_counts.get('Colours', 0)
            total_colours += colour_count

        # calculate percentage
        percent_chemical = (total_colours / total_clients) if total_clients > 0 else 0

        return percent_chemical
        

    def save(self, filename):
        """
        Save the workbook to a file.

        Args:
            filename: Path to save the Excel file
        """
        self.workbook.save(filename)
        print(f"Report saved to {filename}")
